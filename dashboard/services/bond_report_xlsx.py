"""
Official Surety Bond & Discharge Reports (XLSX)
================================================
Modern Fortune-50-style internal reports modeled on Shamrock's long-running
OSI / Palmetto bond report workbooks:

  Header:
    Shamrock Bail Bonds - 1528 Broadway Ft. Myers, FL 33901
    Owner: Brendan O'Neal          Report Date: MM/DD/YYYY
    Surety: OSI | PALMETTO         Report Type: Active Bonds | Discharges | …

  Columns (active / liability report):
    Count | Power # | Defendant First Name | Defendant's Last Name |
    Bond Date | Bond Liability | Gross Premium | Premium | Buf | Collateral

  Extra sheets:
    Expired and Voided Powers | Transfers | Discharges (when requested)

OSI rates (per $100 premium face): surety $7.50, BUF $5.00
Palmetto rates: surety $10.00, BUF $5.00
"""
from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from typing import Any, Iterable, Optional

logger = logging.getLogger(__name__)

AGENCY_NAME = "Shamrock Bail Bonds"
AGENCY_ADDRESS = "1528 Broadway Ft. Myers, FL 33901"
OWNER_NAME = "Brendan O'Neal"
AGENT_LICENSE = "P139768"

# Brand greens (official internal report palette)
BRAND_DARK = "0B3D2E"
BRAND_MID = "1A5F2A"
BRAND_ACCENT = "00A86B"
HEADER_BG = "0F172A"
HEADER_FG = "FFFFFF"
ALT_ROW = "F0FDF4"
TOTAL_BG = "ECFDF5"
BORDER = "CBD5E1"
VOID_BG = "FEF2F2"
DISCHARGE_BG = "EFF6FF"

SURETY_RATES = {
    "OSI": {"surety_per_100": 7.50, "buf_per_100": 5.00, "label": "OSI / Bankers"},
    "PALMETTO": {"surety_per_100": 10.00, "buf_per_100": 5.00, "label": "Palmetto Surety"},
}

ACTIVE_HEADERS = [
    "Count",
    "Power #",
    "Defendant First Name",
    "Defendant's Last Name",
    "Bond Date",
    "Bond Liability",
    "Gross Premium",
    "Premium",  # surety share
    "Buf",
    "Collateral",
    "County",
    "Status",
    "Case Number",
]

DISCHARGE_HEADERS = [
    "Count",
    "Power #",
    "Defendant First Name",
    "Defendant's Last Name",
    "Bond Date",
    "Discharge Date",
    "Bond Liability",
    "Gross Premium",
    "Premium",
    "Buf",
    "Discharge Type",
    "County",
    "Case Number",
    "Notes",
]

VOID_HEADERS = ["Count", "Power #", "Status", "Bond Date", "Notes"]


def _require_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
        from openpyxl.utils import get_column_letter
        return Workbook, Alignment, Border, Font, PatternFill, Side, get_column_letter
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is required for official bond reports. "
            "Install with: pip install openpyxl"
        ) from e


def _norm_surety(surety: str | None) -> str:
    s = (surety or "OSI").strip().upper()
    if s in ("PALMETTO", "PSC", "PALMETTO SURETY"):
        return "PALMETTO"
    if s in ("OSI", "BANKERS", "BANKERS INSURANCE"):
        return "OSI"
    if "PALMETTO" in s or s.startswith("PSC"):
        return "PALMETTO"
    return "OSI"


def split_premium(bond_amount: float, surety: str) -> dict[str, float]:
    """Match historical workbook math: gross = max(100, 10% liability); surety/BUF per $100 of gross."""
    ba = float(bond_amount or 0)
    rates = SURETY_RATES[_norm_surety(surety)]
    gross = max(100.0, ba * 0.10) if ba > 0 else 0.0
    # Historical sheets treat "Premium" as surety share = gross * (rate/100)
    # Actually looking at data: liability 5000, gross 500, premium 37.5, buf 25
    # 37.5 = 500 * 0.075, 25 = 500 * 0.05 → rates are % of GROSS PREMIUM
    surety_share = round(gross * (rates["surety_per_100"] / 100.0), 2)
    buf = round(gross * (rates["buf_per_100"] / 100.0), 2)
    return {
        "bond_liability": round(ba, 2),
        "gross_premium": round(gross, 2),
        "premium": surety_share,
        "buf": buf,
    }


def _split_name(full: str | None, first: str | None = None, last: str | None = None) -> tuple[str, str]:
    if first or last:
        return (first or "").strip(), (last or "").strip()
    name = (full or "").strip()
    if not name:
        return "", ""
    if "," in name:
        parts = [p.strip() for p in name.split(",", 1)]
        return (parts[1] if len(parts) > 1 else ""), parts[0]
    bits = name.split()
    if len(bits) == 1:
        return bits[0], ""
    return bits[0], " ".join(bits[1:])


def _fmt_date(val: Any) -> Optional[datetime]:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.replace(tzinfo=None) if val.tzinfo else val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s[:19].replace("Z", ""), fmt.replace("T%H:%M:%S", "") if "T" not in s else fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).replace(tzinfo=None)
    except Exception:
        return None


def _money(n: float) -> float:
    return round(float(n or 0), 2)


def _style_workbook(wb, sheet_title: str = "Report"):
    Workbook, Alignment, Border, Font, PatternFill, Side, get_column_letter = _require_openpyxl()
    # styles applied in build functions
    return Alignment, Border, Font, PatternFill, Side, get_column_letter


def build_official_bond_report(
    bonds: Iterable[dict],
    *,
    surety: str = "OSI",
    report_type: str = "Active Bonds",
    report_date: datetime | None = None,
    voids: Iterable[dict] | None = None,
    discharges: Iterable[dict] | None = None,
    transfers: Iterable[dict] | None = None,
    title_override: str | None = None,
) -> bytes:
    """
    Build a multi-sheet official workbook. Returns .xlsx bytes.
    """
    Workbook, Alignment, Border, Font, PatternFill, Side, get_column_letter = _require_openpyxl()

    surety_key = _norm_surety(surety)
    report_date = report_date or datetime.now(timezone.utc).replace(tzinfo=None)
    bonds = list(bonds)
    voids = list(voids or [])
    discharges = list(discharges or [])
    transfers = list(transfers or [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Report" if surety_key == "PALMETTO" else "Active Bonds"

    thin = Border(
        left=Side(style="thin", color=BORDER),
        right=Side(style="thin", color=BORDER),
        top=Side(style="thin", color=BORDER),
        bottom=Side(style="thin", color=BORDER),
    )
    fill_header = PatternFill("solid", fgColor=HEADER_BG)
    fill_brand = PatternFill("solid", fgColor=BRAND_DARK)
    fill_alt = PatternFill("solid", fgColor=ALT_ROW)
    fill_total = PatternFill("solid", fgColor=TOTAL_BG)
    font_title = Font(name="Calibri", size=16, bold=True, color=HEADER_FG)
    font_sub = Font(name="Calibri", size=11, color="E2E8F0")
    font_col = Font(name="Calibri", size=10, bold=True, color=HEADER_FG)
    font_cell = Font(name="Calibri", size=10, color="0F172A")
    font_total = Font(name="Calibri", size=10, bold=True, color=BRAND_DARK)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # ── Title block ──
    ws.merge_cells("A1:M1")
    ws["A1"] = f"{AGENCY_NAME} — {AGENCY_ADDRESS}"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_brand
    ws["A1"].alignment = left

    ws.merge_cells("A2:M2")
    rates = SURETY_RATES[surety_key]
    ws["A2"] = (
        f"Owner: {OWNER_NAME}  ·  Agent Lic. #{AGENT_LICENSE}  ·  "
        f"Surety: {rates['label']}  ·  {title_override or report_type}  ·  "
        f"Report Date: {report_date.strftime('%m/%d/%Y')}"
    )
    ws["A2"].font = font_sub
    ws["A2"].fill = PatternFill("solid", fgColor=BRAND_MID)
    ws["A2"].alignment = left

    ws.merge_cells("A3:M3")
    ws["A3"] = (
        f"CONFIDENTIAL — Internal surety production / liability report. "
        f"Generated automatically by Shamrock Super CRM. "
        f"Premium rates: Surety ${rates['surety_per_100']:.2f} / BUF ${rates['buf_per_100']:.2f} per $100 gross premium."
    )
    ws["A3"].font = Font(name="Calibri", size=9, italic=True, color="64748B")

    # ── Column headers row 5 ──
    header_row = 5
    for col, h in enumerate(ACTIVE_HEADERS, 1):
        cell = ws.cell(header_row, col, h)
        cell.font = font_col
        cell.fill = fill_header
        cell.alignment = center
        cell.border = thin

    # ── Data ──
    total_liability = 0.0
    total_gross = 0.0
    total_premium = 0.0
    total_buf = 0.0
    row_i = header_row + 1
    count = 0

    for b in bonds:
        # Filter by surety when field present
        b_surety = _norm_surety(
            b.get("surety") or b.get("insurance_company") or b.get("surety_id") or surety_key
        )
        if b_surety != surety_key and (b.get("surety") or b.get("surety_id") or b.get("insurance_company")):
            continue

        status = str(b.get("status") or "active").lower()
        if status in ("void", "voided", "expired"):
            continue

        count += 1
        first, last = _split_name(
            b.get("defendant_name") or b.get("full_name") or b.get("name"),
            b.get("defendant_first_name") or b.get("first_name"),
            b.get("defendant_last_name") or b.get("last_name"),
        )
        ba = float(b.get("bond_amount") or b.get("bond_liability") or b.get("amount") or 0)
        money = split_premium(ba, surety_key)
        # Allow explicit premium override from record
        if b.get("gross_premium") is not None:
            money["gross_premium"] = _money(b.get("gross_premium"))
        if b.get("premium") is not None and b.get("buf") is not None:
            money["premium"] = _money(b.get("premium"))
            money["buf"] = _money(b.get("buf"))

        power = (
            b.get("power_number")
            or b.get("poa_number")
            or b.get("power")
            or b.get("power_no")
            or ""
        )
        collateral = (
            b.get("collateral")
            or b.get("collateral_description")
            or ("Promissory Note, Indemnity Agreement" if surety_key == "OSI" else "Indem Agreement/Prom. Note")
        )
        bond_dt = _fmt_date(b.get("bond_date") or b.get("posted_date") or b.get("created_at"))

        values = [
            count,
            power,
            first,
            last,
            bond_dt,
            money["bond_liability"],
            money["gross_premium"],
            money["premium"],
            money["buf"],
            collateral,
            b.get("county") or "",
            (b.get("status") or "active").title(),
            b.get("case_number") or b.get("case_no") or "",
        ]

        total_liability += money["bond_liability"]
        total_gross += money["gross_premium"]
        total_premium += money["premium"]
        total_buf += money["buf"]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row_i, col, val)
            cell.font = font_cell
            cell.border = thin
            cell.alignment = center if col in (1, 5, 11, 12) else (right if col in (6, 7, 8, 9) else left)
            if row_i % 2 == 0:
                cell.fill = fill_alt
            if col in (6, 7, 8, 9) and isinstance(val, (int, float)):
                cell.number_format = '"$"#,##0.00'
            if col == 5 and isinstance(val, datetime):
                cell.number_format = "MM/DD/YYYY"
        row_i += 1

    # Totals
    ws.cell(row_i, 1, "")
    ws.cell(row_i, 4, "TOTALS").font = font_total
    for col, val in ((6, total_liability), (7, total_gross), (8, total_premium), (9, total_buf)):
        cell = ws.cell(row_i, col, val)
        cell.font = font_total
        cell.fill = fill_total
        cell.border = thin
        cell.number_format = '"$"#,##0.00'
    for col in range(1, 14):
        c = ws.cell(row_i, col)
        c.fill = fill_total
        c.border = thin

    row_i += 2
    ws.cell(row_i, 1, f"Line items: {count}  ·  Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
    ws.cell(row_i, 1).font = Font(name="Calibri", size=9, italic=True, color="64748B")

    # Column widths
    widths = [8, 18, 16, 18, 12, 14, 14, 12, 10, 32, 12, 12, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A6"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # ── Voids sheet ──
    ws_void = wb.create_sheet("Expired and Voided Powers")
    ws_void["A1"] = f"{AGENCY_NAME} — Voided / Expired Powers ({surety_key})"
    ws_void["A1"].font = font_title
    ws_void["A1"].fill = fill_brand
    ws_void.merge_cells("A1:D1")
    for col, h in enumerate(VOID_HEADERS, 1):
        cell = ws_void.cell(3, col, h)
        cell.font = font_col
        cell.fill = fill_header
        cell.border = thin
    vr = 4
    for i, v in enumerate(voids, 1):
        power = v.get("power_number") or v.get("poa_number") or v.get("power") or ""
        row_vals = [i, power, (v.get("status") or "VOID").upper(), _fmt_date(v.get("bond_date")), v.get("notes") or "VOID"]
        for col, val in enumerate(row_vals, 1):
            cell = ws_void.cell(vr, col, val)
            cell.border = thin
            cell.fill = PatternFill("solid", fgColor=VOID_BG)
            if col == 4 and isinstance(val, datetime):
                cell.number_format = "MM/DD/YYYY"
        vr += 1
    for i, w in enumerate([8, 20, 12, 14, 24], 1):
        ws_void.column_dimensions[get_column_letter(i)].width = w

    # ── Discharges sheet ──
    ws_dis = wb.create_sheet("Discharges")
    ws_dis["A1"] = f"{AGENCY_NAME} — Bond Discharges / Exonerations ({surety_key})"
    ws_dis["A1"].font = font_title
    ws_dis["A1"].fill = fill_brand
    ws_dis.merge_cells("A1:N1")
    ws_dis["A2"] = (
        f"Owner: {OWNER_NAME}  ·  Report Date: {report_date.strftime('%m/%d/%Y')}  ·  "
        "Official discharge register (same column grammar as bond liability report)"
    )
    ws_dis["A2"].font = font_sub
    ws_dis["A2"].fill = PatternFill("solid", fgColor=BRAND_MID)
    ws_dis.merge_cells("A2:N2")
    for col, h in enumerate(DISCHARGE_HEADERS, 1):
        cell = ws_dis.cell(4, col, h)
        cell.font = font_col
        cell.fill = fill_header
        cell.border = thin
        cell.alignment = center

    dr = 5
    d_count = 0
    d_liab = d_gross = d_prem = d_buf = 0.0
    for d in discharges:
        d_surety = _norm_surety(d.get("surety") or d.get("surety_id") or d.get("insurance_company") or surety_key)
        if d_surety != surety_key and (d.get("surety") or d.get("surety_id")):
            continue
        d_count += 1
        first, last = _split_name(
            d.get("defendant_name") or d.get("full_name"),
            d.get("defendant_first_name") or d.get("first_name"),
            d.get("defendant_last_name") or d.get("last_name"),
        )
        ba = float(d.get("bond_amount") or d.get("bond_liability") or 0)
        money = split_premium(ba, surety_key)
        power = d.get("power_number") or d.get("poa_number") or d.get("power") or ""
        vals = [
            d_count,
            power,
            first,
            last,
            _fmt_date(d.get("bond_date")),
            _fmt_date(d.get("discharge_date") or d.get("exonerated_at") or d.get("updated_at")),
            money["bond_liability"],
            money["gross_premium"],
            money["premium"],
            money["buf"],
            (d.get("status") or d.get("discharge_type") or "exonerated").title(),
            d.get("county") or "",
            d.get("case_number") or "",
            d.get("notes") or d.get("discharge_note") or "",
        ]
        d_liab += money["bond_liability"]
        d_gross += money["gross_premium"]
        d_prem += money["premium"]
        d_buf += money["buf"]
        for col, val in enumerate(vals, 1):
            cell = ws_dis.cell(dr, col, val)
            cell.border = thin
            cell.font = font_cell
            if dr % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=DISCHARGE_BG)
            if col in (7, 8, 9, 10) and isinstance(val, (int, float)):
                cell.number_format = '"$"#,##0.00'
            if col in (5, 6) and isinstance(val, datetime):
                cell.number_format = "MM/DD/YYYY"
        dr += 1

    for col, val in ((7, d_liab), (8, d_gross), (9, d_prem), (10, d_buf)):
        cell = ws_dis.cell(dr, col, val)
        cell.font = font_total
        cell.fill = fill_total
        cell.number_format = '"$"#,##0.00'
        cell.border = thin
    ws_dis.cell(dr, 4, "TOTALS").font = font_total
    for i, w in enumerate([8, 18, 14, 16, 12, 14, 14, 12, 12, 10, 14, 12, 14, 28], 1):
        ws_dis.column_dimensions[get_column_letter(i)].width = w

    # ── Transfers sheet (Palmetto layout compatibility) ──
    ws_tr = wb.create_sheet("Transfers")
    ws_tr["A1"] = f"{AGENCY_NAME} — Transfers ({surety_key})"
    ws_tr["A1"].fill = fill_brand
    ws_tr["A1"].font = font_title
    tr_headers = ["Count", "Power #", "Defendant First Name", "Defendant's Last Name", "Bond Date", "Bond Liability", "Agency"]
    for col, h in enumerate(tr_headers, 1):
        cell = ws_tr.cell(3, col, h)
        cell.font = font_col
        cell.fill = fill_header
        cell.border = thin
    tr = 4
    for i, t in enumerate(transfers, 1):
        first, last = _split_name(t.get("defendant_name"), t.get("first_name"), t.get("last_name"))
        vals = [
            i,
            t.get("power_number") or t.get("power") or "",
            first,
            last,
            _fmt_date(t.get("bond_date")),
            float(t.get("bond_amount") or 0),
            t.get("agency") or AGENCY_NAME,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws_tr.cell(tr, col, val)
            cell.border = thin
            if col == 6 and isinstance(val, (int, float)):
                cell.number_format = '"$"#,##0.00'
            if col == 5 and isinstance(val, datetime):
                cell.number_format = "MM/DD/YYYY"
        tr += 1
    for i, w in enumerate([8, 18, 16, 18, 12, 14, 22], 1):
        ws_tr.column_dimensions[get_column_letter(i)].width = w

    # ── Summary cover ──
    ws_sum = wb.create_sheet("Executive Summary", 0)
    ws_sum["A1"] = f"{AGENCY_NAME} — Executive Summary"
    ws_sum["A1"].font = font_title
    ws_sum["A1"].fill = fill_brand
    ws_sum.merge_cells("A1:B1")
    summary_rows = [
        ("Surety", rates["label"]),
        ("Report type", title_override or report_type),
        ("Report date", report_date.strftime("%m/%d/%Y")),
        ("Active line items", count),
        ("Total bond liability", f"${total_liability:,.2f}"),
        ("Total gross premium", f"${total_gross:,.2f}"),
        ("Total surety premium", f"${total_premium:,.2f}"),
        ("Total BUF", f"${total_buf:,.2f}"),
        ("Voided powers listed", len(voids)),
        ("Discharges listed", d_count),
        ("Transfers listed", len(transfers)),
        ("Prepared for", "Internal / Surety submission"),
        ("Prepared by", f"{OWNER_NAME} (#{AGENT_LICENSE})"),
        ("System", "Shamrock Super CRM — Official Report Engine"),
    ]
    for i, (k, v) in enumerate(summary_rows, 3):
        ws_sum.cell(i, 1, k).font = Font(name="Calibri", bold=True, size=11)
        ws_sum.cell(i, 2, v).font = Font(name="Calibri", size=11)
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 42

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def filename_for(surety: str, report_type: str = "Bond Report", when: datetime | None = None) -> str:
    when = when or datetime.now(timezone.utc)
    s = _norm_surety(surety)
    safe = report_type.replace(" ", "_")
    return f"Shamrock_{s}_{safe}_{when.strftime('%Y-%m-%d')}.xlsx"
